"""
اولویت‌بندی صفحات از خروجی Performance GSC.
"""
from __future__ import annotations

from pathlib import Path
from urllib.parse import unquote, urlparse

_GSC_DIR = Path(__file__).resolve().parents[1] / "Seo_search console"
_PERF_FILE = "saroshan.ir-Performance-on-Search-2026-06-25.xlsx"


def _slug_from_url(url: str) -> tuple[str, str] | None:
    path = unquote(urlparse(str(url)).path.strip("/"))
    if path.startswith("رشته/"):
        return "major", path.split("/", 1)[1]
    if path.startswith("دانشگاه/"):
        return "university", path.split("/", 1)[1]
    if path.startswith("blog/"):
        return "blog", path.split("/", 1)[1].strip("/")
    if path.startswith("کشور/"):
        return "country", path.split("/", 1)[1].strip("/")
    return None


def _load_pages_from_analysis_json() -> dict[str, set[str]]:
    """Fallback when xlsx folder is missing."""
    analysis = Path(__file__).resolve().parents[1] / "scripts" / "gsc_analysis_output.json"
    empty = {"majors": set(), "universities": set(), "blogs": set(), "countries": set()}
    if not analysis.is_file():
        return empty
    import json

    data = json.loads(analysis.read_text(encoding="utf-8"))
    majors: set[str] = set()
    universities: set[str] = set()
    blogs: set[str] = set()
    countries: set[str] = set()
    for key, blob in data.items():
        if "Performance" not in key:
            continue
        for row in blob.get("sheets", {}).get("Pages", {}).get("rows", []):
            if not row or not isinstance(row[0], str) or not row[0].startswith("http"):
                continue
            parsed = _slug_from_url(row[0])
            if not parsed:
                continue
            kind, slug = parsed
            if kind == "major":
                majors.add(slug)
            elif kind == "university":
                universities.add(slug)
            elif kind == "blog":
                blogs.add(slug)
            elif kind == "country":
                countries.add(slug)
        break
    return {
        "majors": majors,
        "universities": universities,
        "blogs": blogs,
        "countries": countries,
    }


def load_gsc_performance_priorities(
    *,
    min_impressions: int = 50,
    low_ctr_threshold: float = 0.04,
    top_n: int = 80,
) -> dict[str, set[str]]:
    """
    فاز ۱: صفحات پربازدید + CTR پایین (فرصت بهبود محتوا).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl required") from exc

    path = _GSC_DIR / _PERF_FILE
    if not path.is_file():
        candidates = sorted(_GSC_DIR.glob("*Performance*.xlsx"), reverse=True)
        path = candidates[0] if candidates else path

    if not path.is_file():
        return _load_pages_from_analysis_json()

    majors: set[str] = set()
    universities: set[str] = set()
    blogs: set[str] = set()
    countries: set[str] = set()

    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["Pages"].iter_rows(min_row=2, values_only=True))
    wb.close()

    scored: list[tuple[float, str, str]] = []
    for row in rows:
        if not row or not row[0]:
            continue
        url = str(row[0]).strip()
        if not url.startswith("http"):
            continue
        clicks = float(row[1] or 0)
        impressions = float(row[2] or 0)
        ctr = float(row[3]) if row[3] not in ("", None) else 0.0
        if impressions < min_impressions:
            continue
        parsed = _slug_from_url(url)
        if not parsed:
            continue
        kind, slug = parsed
        # امتیاز: نمایش بالا + CTR پایین = اولویت بهبود
        score = impressions * (1.0 if ctr < low_ctr_threshold else 0.35) + clicks * 2
        scored.append((score, kind, slug))

    scored.sort(reverse=True)
    for _score, kind, slug in scored[:top_n]:
        if kind == "major":
            majors.add(slug)
        elif kind == "university":
            universities.add(slug)
        elif kind == "blog":
            blogs.add(slug)
        elif kind == "country":
            countries.add(slug)

    return {
        "majors": majors,
        "universities": universities,
        "blogs": blogs,
        "countries": countries,
    }
