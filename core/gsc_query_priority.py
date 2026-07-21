"""
اولویت محتوا بر اساس جستجوهای واقعی کاربران در GSC (Queries).
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

_GSC_DIR = Path(__file__).resolve().parents[1] / "Seo_search console"
_ANALYSIS_JSON = Path(__file__).resolve().parents[1] / "scripts" / "gsc_analysis_output.json"
_CACHE_PATH = (
    Path(__file__).resolve().parent / "seed_data" / "cache" / "gsc_search_queries.json"
)

# کلیدواژه فارسی/انگلیسی → اسلاگ دانشگاه
_UNIVERSITY_QUERY_RULES: tuple[tuple[tuple[str, ...], str], ...] = (
    (("پکن", "peking"), "peking-university"),
    (("فودان", "fudan"), "fudan-university"),
    (("کمپلوتنسه", "complutense"), "universidad-complutense-madrid"),
    (("والنسیا", "valencia"), "universitat-de-valencia"),
    (("سیچوان", "sichuan"), "sichuan-university"),
    (("هونان", "hunan"), "hunan-university"),
    (("سون یات", "sun yat", "zhongshan"), "sun-yat-sen-university"),
    (("ویندزور", "windsor"), "university-of-windsor"),
    (("گرانادا", "granada"), "universidad-de-granada"),
    (("مموریال", "memorial"), "memorial-university"),
    (("بیهانگ", "beihang"), "beihang-university"),
    (("یورک", "york university"), "york-university"),
    (("کنکوردیا", "concordia"), "concordia-university"),
    (("آلبرتا", "alberta"), "university-of-alberta"),
    (("شیمین", "xiamen"), "xiamen-university"),
    (("بوئنوس آیرس", "buenos aires"), "universidad-de-buenos-aires"),
    (("شاندونگ", "shandong"), "shandong-university"),
    (("آنهویی", "anhui"), "anhui-university"),
    (("فوردهام", "fordham"), "fordham-university"),
)

# کلیدواژه → (کشور, عبارت رشته برای جستجو در title)
_MAJOR_QUERY_RULES: tuple[tuple[tuple[str, ...], str, str], ...] = (
    (("فناوری اطلاعات سلامت", "health information"), "australia", "فناوری اطلاعات سلامت"),
    (("مهندسی مواد", "متالورژی"), "canada", "مهندسی مواد"),
    (("پزشکی", "تحصیل پزشکی"), "japan", "پزشکی"),
    (("دندانپزشکی",), "india", "دندانپزشکی"),
    (("طراحی لباس",), "south_korea", "طراحی لباس"),
    (("حسابداری",), "south_korea", "حسابداری"),
    (("فیزیوتراپی",), "south_korea", "فیزیوتراپی"),
    (("نوروساینس",), "canada", "علوم اعصاب"),
    (("جرم شناسی", "جرم‌شناسی"), "canada", "جرم"),
    (("ادبیات فارسی",), "canada", "ادبیات"),
    (("هوش مصنوعی",), "canada", "هوش مصنوعی"),
    (("موسیقی",), "germany", "موسیقی"),
)

_COUNTRY_QUERY_KEYWORDS: dict[str, tuple[str, ...]] = {
    "china": ("چین", "csc", "یوان", "پکن", "شانگهای", "hsk", "x1", "x2"),
    "germany": ("آلمان", "تمکن مالی", "blocked", "آلمانی"),
    "spain": ("اسپانیا", "tapas", "تاپاس", "ویزای d"),
    "canada": ("کانادا", "pgwp", "ielts"),
    "france": ("فرانسه", "campus france", "تمکن"),
    "south_korea": ("کره", "topik", "gks"),
    "japan": ("ژاپن", "mext", "jlpt"),
}


def _normalize_query(q: str) -> str:
    return re.sub(r"\s+", " ", (q or "").strip().lower())


def _load_query_rows() -> list[tuple[str, float, float]]:
    """(query, clicks, impressions) از xlsx یا JSON تحلیل."""
    rows: list[tuple[str, float, float]] = []

    xlsx_candidates = sorted(_GSC_DIR.glob("*Performance*.xlsx"), reverse=True)
    if xlsx_candidates:
        try:
            from openpyxl import load_workbook
        except ImportError:
            xlsx_candidates = []

    for path in xlsx_candidates[:1]:
        wb = load_workbook(path, read_only=True, data_only=True)
        if "Queries" not in wb.sheetnames:
            wb.close()
            continue
        ws = wb["Queries"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            q = str(row[0]).strip()
            if not q or q.lower().startswith("top quer"):
                continue
            rows.append((q, float(row[1] or 0), float(row[2] or 0)))
        wb.close()
        if rows:
            return rows

    if _ANALYSIS_JSON.is_file():
        data = json.loads(_ANALYSIS_JSON.read_text(encoding="utf-8"))
        for key, blob in data.items():
            if "Performance" not in key:
                continue
            for row in blob.get("sheets", {}).get("Queries", {}).get("rows", []):
                if not row or not isinstance(row[0], str):
                    continue
                q = row[0].strip()
                if q.lower().startswith("top quer"):
                    continue
                rows.append((q, float(row[1] or 0), float(row[2] or 0)))
            break

    if _CACHE_PATH.is_file() and not rows:
        cached = json.loads(_CACHE_PATH.read_text(encoding="utf-8"))
        for item in cached.get("queries", []):
            rows.append((item["q"], float(item.get("clicks", 0)), float(item.get("impressions", 0))))

    return rows


def write_gsc_queries_cache(rows: list[tuple[str, float, float]] | None = None) -> Path:
    if rows is None:
        rows = _load_query_rows()
    rows_sorted = sorted(rows, key=lambda x: x[2], reverse=True)
    payload = {
        "queries": [
            {"q": q, "clicks": c, "impressions": i}
            for q, c, i in rows_sorted
        ]
    }
    _CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    _CACHE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return _CACHE_PATH


def _match_university_slug(query: str) -> str | None:
    nq = _normalize_query(query)
    for keys, slug in _UNIVERSITY_QUERY_RULES:
        if any(k in nq for k in keys):
            return slug
    return None


def _match_major_slug(query: str) -> str | None:
    from core.models import Major

    nq = _normalize_query(query)
    for keys, country, title_hint in _MAJOR_QUERY_RULES:
        if any(k in nq for k in keys):
            major = (
                Major.objects.filter(is_active=True, country=country, title__icontains=title_hint)
                .order_by("order", "id")
                .first()
            )
            if major:
                return major.slug
    if "رشته" in nq or "تحصیل" in nq:
        for country, country_keys in _COUNTRY_QUERY_KEYWORDS.items():
            if any(k in nq for k in country_keys):
                token = nq
                for w in ("تحصیل", "در", "رشته", "بورسیه", "مهاجرت", country, "چین", "ژاپن"):
                    token = token.replace(w, " ")
                token = token.strip()
                if len(token) >= 3:
                    major = (
                        Major.objects.filter(is_active=True, country=country, title__icontains=token[:20])
                        .first()
                    )
                    if major:
                        return major.slug
    return None


def resolve_queries_to_slugs(
    *,
    min_impressions: int = 3,
    top_n: int = 120,
) -> dict[str, Any]:
    """
    نگاشت جستجوها به اسلاگ — خروجی برای refresh_rich_content.

    Returns:
        majors, universities: set[str]
        queries_by_major, queries_by_university: dict[slug, list[str]]
        top_queries: list[dict]
    """
    rows = _load_query_rows()
    rows = sorted(rows, key=lambda x: x[2], reverse=True)
    if not rows and _CACHE_PATH.is_file():
        write_gsc_queries_cache()

    majors: set[str] = set()
    universities: set[str] = set()
    queries_by_major: dict[str, list[str]] = {}
    queries_by_university: dict[str, list[str]] = {}
    top_queries: list[dict] = []

    for query, clicks, impressions in rows[:top_n]:
        if impressions < min_impressions:
            continue
        top_queries.append(
            {"q": query, "clicks": clicks, "impressions": impressions}
        )
        uni_slug = _match_university_slug(query)
        if uni_slug:
            universities.add(uni_slug)
            queries_by_university.setdefault(uni_slug, []).append(query)
            continue
        major_slug = _match_major_slug(query)
        if major_slug:
            majors.add(major_slug)
            queries_by_major.setdefault(major_slug, []).append(query)

    return {
        "majors": majors,
        "universities": universities,
        "queries_by_major": queries_by_major,
        "queries_by_university": queries_by_university,
        "top_queries": top_queries,
    }


def load_query_priority_slugs(*, min_impressions: int = 3, top_n: int = 120) -> dict[str, set[str]]:
    data = resolve_queries_to_slugs(min_impressions=min_impressions, top_n=top_n)
    return {
        "majors": set(data["majors"]),
        "universities": set(data["universities"]),
    }
