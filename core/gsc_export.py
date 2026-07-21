"""
خواندن خروجی Google Search Console (xlsx) — شیت Table.
"""
from __future__ import annotations

import json
from pathlib import Path
from urllib.parse import unquote, urlparse

_GSC_DIR = Path(__file__).resolve().parents[1] / "Seo_search console"
_CRAWLED_FILE = "Crawled - currently not indexed.xlsx"
_DISCOVERED_FILE = "Discovered - currently not indexed.xlsx"
_NOT_INDEXED_FILES = (_CRAWLED_FILE, _DISCOVERED_FILE)
_INDEX_CACHE_PATH = (
    Path(__file__).resolve().parent / "seed_data" / "cache" / "gsc_not_indexed_slugs.json"
)


def slug_from_gsc_url(url: str) -> tuple[str, str] | None:
    path = unquote(urlparse(url).path.strip("/"))
    if path.startswith("دانشگاه/"):
        return "university", path.split("/", 1)[1].rstrip("/")
    if path.startswith("رشته/"):
        return "major", path.split("/", 1)[1].rstrip("/")
    if path.startswith("blog/"):
        return "blog", path.split("/", 1)[1].rstrip("/")
    return None


def _load_slugs_from_files(
  names: tuple[str, ...],
  *,
  gsc_dir: Path | None = None,
) -> dict[str, set[str]]:
    """استخراج اسلاگ از یک یا چند فایل GSC."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc

    base = gsc_dir or _GSC_DIR
    majors: set[str] = set()
    universities: set[str] = set()
    blogs: set[str] = set()

    for name in names:
        path = base / name
        if not path.is_file():
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["Table"] if "Table" in wb.sheetnames else wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            url = str(row[0] or "").strip()
            if not url.startswith("http"):
                continue
            parsed = slug_from_gsc_url(url)
            if not parsed:
                continue
            kind, slug = parsed
            if kind == "major":
                majors.add(slug)
            elif kind == "university":
                universities.add(slug)
            else:
                blogs.add(slug)
        wb.close()

    return {"majors": majors, "universities": universities, "blogs": blogs}


def load_gsc_not_indexed_slugs(
    *,
    gsc_dir: Path | None = None,
) -> dict[str, set[str]]:
    """هر دو گزارش Crawled + Discovered."""
    return _load_slugs_from_files(_NOT_INDEXED_FILES, gsc_dir=gsc_dir)


def load_gsc_crawled_not_indexed_slugs(
    *,
    gsc_dir: Path | None = None,
) -> dict[str, set[str]]:
    """فقط Crawled - currently not indexed."""
    return _load_slugs_from_files((_CRAWLED_FILE,), gsc_dir=gsc_dir)


def load_gsc_discovered_not_indexed_slugs(
    *,
    gsc_dir: Path | None = None,
) -> dict[str, set[str]]:
    """فقط Discovered - currently not indexed."""
    return _load_slugs_from_files((_DISCOVERED_FILE,), gsc_dir=gsc_dir)


def write_gsc_not_indexed_cache(
    slugs: dict[str, set[str]] | None = None,
    *,
    path: Path | None = None,
) -> Path:
    """ذخیرهٔ اسلاگ‌ها برای استفاده در runtime (لینک داخلی / sitemap)."""
    if slugs is None:
        slugs = load_gsc_not_indexed_slugs()
    target = path or _INDEX_CACHE_PATH
    target.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "majors": sorted(slugs["majors"]),
        "universities": sorted(slugs["universities"]),
        "blogs": sorted(slugs["blogs"]),
    }
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return target
