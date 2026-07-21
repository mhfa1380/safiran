"""وارد کردن درخواست‌های مشاوره/ارزیابی از خروجی Excel ادمین (django-import-export)."""

from __future__ import annotations

import json
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db.models.signals import post_save
from django.utils import timezone

from core.evaluation_share import create_evaluation_share
from core.models import ConsultationRequest, ConsultationSlot, EvaluationRequest
from core.signals import (
    notify_new_consultation_request,
    notify_new_evaluation_request,
    persist_evaluation_snapshot,
)

CONSULTATION_COLUMNS = {
    "نام و نام خانوادگی": "full_name",
    "شماره تماس": "phone",
    "ایمیل": "email",
    "نوع مشاوره": "consultation_type",
    "کشور مقصد": "country",
    "زمان رزرو": "slot_id",
    "دانشگاه مورد علاقه": "interest_university_id",
    "توضیحات و سوالات شما": "description",
    "از کجا با ما آشنا شدید؟": "referral_source",
    "شبکه اجتماعی": "referral_social_platform",
    "جزئیات منبع": "referral_detail",
    "وضعیت درخواست": "status",
    "مشاهده توسط ادمین": "admin_seen_at",
    "تاریخ ثبت": "created_at",
    "آخرین به‌روزرسانی": "updated_at",
}

EVALUATION_COLUMNS = {
    "نام و نام خانوادگی": "full_name",
    "شماره تماس": "phone",
    "ایمیل": "email",
    "سال تولد": "birth_year",
    "وضعیت تاهل": "marital_status",
    "کی قصد اپلای دارید؟": "apply_timeline",
    "تمکن مالی بالای یک میلیارد تومان دارم": "has_financial_capacity",
    "آخرین مدرک تحصیلی": "current_degree",
    "رشته تحصیلی فعلی / قبلی": "field_of_study",
    "معدل (تقریبی)": "average_grade",
    "سال فارغ التحصیلی": "graduation_year",
    "کشور مقصد": "target_country",
    "نوع آزمون زبان": "language_test_type",
    "مدرک زبان (آیلتس/تافل) دارم": "has_ielts",
    "نمره زبان (در صورت وجود)": "language_score",
    "مقاله ژورنالی": "has_journal_article",
    "مقاله کنفرانسی": "has_conference_article",
    "چاپ یا ترجمه کتاب": "has_book",
    "آزمون‌های بین‌الملل (SAT, GRE, GMAT ...)": "has_international_tests",
    "کشورهای مورد نظر": "desired_countries",
    "رشته مورد نظر (اختیاری)": "desired_major",
    "تمایل دارید چه بخشی از کار به موسسه سپرده شود؟": "service_scope",
    "ترم / سال مورد نظر برای شروع": "preferred_intake",
    "توضیحات تکمیلی": "notes",
    "از کجا با ما آشنا شدید؟": "referral_source",
    "شبکه اجتماعی": "referral_social_platform",
    "جزئیات منبع": "referral_detail",
    "وضعیت پیگیری": "status",
    "اولویت": "priority",
    "نیاز به پیگیری": "follow_up_required",
    "دسته پیگیری": "follow_up_category",
    "نتیجه آخرین تماس": "contact_result",
    "یادداشت کارشناس (داخلی)": "admin_notes",
    "پیشنهاد هوشمند (ذخیره خودکار)": "recommendation_snapshot",
    "زمان آخرین تماس": "contacted_at",
    "زمان پیگیری بعدی": "next_follow_up_at",
    "کارشناس مسئول": "assigned_to_id",
    "مشاهده توسط ادمین": "admin_seen_at",
    "تاریخ ثبت": "created_at",
    "آخرین به‌روزرسانی": "updated_at",
}

BOOL_FIELDS = {
    "has_financial_capacity",
    "has_ielts",
    "has_journal_article",
    "has_conference_article",
    "has_book",
    "has_international_tests",
    "follow_up_required",
}

INT_FIELDS = {"birth_year", "slot_id", "interest_university_id", "assigned_to_id"}

DATETIME_FIELDS = {
    "admin_seen_at",
    "created_at",
    "updated_at",
    "contacted_at",
    "next_follow_up_at",
}

BLANK_OK = {
    "email",
    "description",
    "referral_source",
    "referral_social_platform",
    "referral_detail",
    "graduation_year",
    "language_score",
    "desired_countries",
    "desired_major",
    "service_scope",
    "preferred_intake",
    "notes",
    "follow_up_category",
    "admin_notes",
    "marital_status",
    "apply_timeline",
}


def _as_bool(value) -> bool:
    if value is None or value == "":
        return False
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def _as_int(value):
    if value is None or value == "":
        return None
    return int(value)


def _as_str(value, *, field: str) -> str:
    if value is None:
        return "" if field in BLANK_OK else ""
    text = str(value).strip()
    return text


def _parse_dt(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                dt = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        else:
            raise ValueError(f"unsupported datetime: {value!r}")
    if timezone.is_naive(dt):
        return timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


def _load_rows(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise CommandError("openpyxl is required: pip install openpyxl") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        out.append(
            {
                headers[i]: row[i] if i < len(row) else None
                for i in range(len(headers))
                if headers[i]
            }
        )
    return out


def _map_row(raw: dict, column_map: dict) -> dict:
    data: dict = {}
    for header, field in column_map.items():
        if header not in raw:
            continue
        value = raw[header]
        if field in BOOL_FIELDS:
            data[field] = _as_bool(value)
        elif field in INT_FIELDS:
            data[field] = _as_int(value)
        elif field in DATETIME_FIELDS:
            data[field] = _parse_dt(value)
        elif field == "recommendation_snapshot":
            if value is None or value == "":
                data[field] = None
            elif isinstance(value, dict):
                data[field] = value
            else:
                text = str(value)
                try:
                    data[field] = json.loads(text)
                except json.JSONDecodeError:
                    # Excel truncates cells at 32767 chars — rebuild report after import.
                    data[field] = None
        else:
            data[field] = _as_str(value, field=field)
    return data


@contextmanager
def _mute_import_side_effects(*, build_reports: bool):
    post_save.disconnect(notify_new_consultation_request, sender=ConsultationRequest)
    post_save.disconnect(notify_new_evaluation_request, sender=EvaluationRequest)
    if not build_reports:
        post_save.disconnect(persist_evaluation_snapshot, sender=EvaluationRequest)
    try:
        yield
    finally:
        post_save.connect(notify_new_consultation_request, sender=ConsultationRequest)
        post_save.connect(notify_new_evaluation_request, sender=EvaluationRequest)
        if not build_reports:
            post_save.connect(persist_evaluation_snapshot, sender=EvaluationRequest)


class Command(BaseCommand):
    help = "Import ConsultationRequest / EvaluationRequest rows from admin Excel export"

    def add_arguments(self, parser):
        parser.add_argument(
            "--consultation",
            default="ConsultationRequest-2026-06-26.xlsx",
            help="Path to consultation export xlsx",
        )
        parser.add_argument(
            "--evaluation",
            default="EvaluationRequest-2026-06-26.xlsx",
            help="Path to evaluation export xlsx",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse only; do not write to database",
        )

        parser.add_argument(
            "--build-reports",
            action="store_true",
            help="Build evaluation snapshots/share links during import (slow)",
        )

    def handle(self, *args, **options):
        root = Path(__file__).resolve().parents[3]
        consultation_path = Path(options["consultation"])
        evaluation_path = Path(options["evaluation"])
        if not consultation_path.is_absolute():
            consultation_path = root / consultation_path
        if not evaluation_path.is_absolute():
            evaluation_path = root / evaluation_path

        if not consultation_path.is_file():
            raise CommandError(f"Consultation file not found: {consultation_path}")
        if not evaluation_path.is_file():
            raise CommandError(f"Evaluation file not found: {evaluation_path}")

        dry_run = options["dry_run"]
        build_reports = options["build_reports"]
        created_c = skipped_c = created_e = skipped_e = 0

        with _mute_import_side_effects(build_reports=build_reports):
            created_c, skipped_c = self._import_consultations(
                consultation_path, dry_run=dry_run
            )
            created_e, skipped_e = self._import_evaluations(
                evaluation_path, dry_run=dry_run, build_reports=build_reports
            )

        self.stdout.write(
            self.style.SUCCESS(
                f"Done: consultations created={created_c} skipped={skipped_c}; "
                f"evaluations created={created_e} skipped={skipped_e}"
            )
        )

    def _import_consultations(self, path: Path, *, dry_run: bool) -> tuple[int, int]:
        created = skipped = 0
        for raw in _load_rows(path):
            phone = str(raw.get("شماره تماس") or "").strip()
            if not phone:
                skipped += 1
                continue
            if ConsultationRequest.objects.filter(phone=phone).exists():
                self.stdout.write(self.style.WARNING(f"Skip consultation phone={phone}"))
                skipped += 1
                continue

            payload = _map_row(raw, CONSULTATION_COLUMNS)
            timestamps = {
                k: payload.pop(k)
                for k in ("created_at", "updated_at", "admin_seen_at")
                if k in payload
            }
            slot_id = payload.pop("slot_id", None)
            university_id = payload.pop("interest_university_id", None)

            if dry_run:
                self.stdout.write(f"Would create consultation {payload.get('full_name')} ({phone})")
                created += 1
                continue

            obj = ConsultationRequest(
                **payload,
                slot_id=slot_id if slot_id else None,
                interest_university_id=university_id if university_id else None,
            )
            obj.save()
            update_fields = {k: v for k, v in timestamps.items() if v is not None}
            if update_fields:
                ConsultationRequest.objects.filter(pk=obj.pk).update(**update_fields)
            if slot_id:
                ConsultationSlot.objects.filter(pk=slot_id).update(is_booked=True)
            created += 1
            self.stdout.write(self.style.SUCCESS(f"Consultation pk={obj.pk} phone={phone}"))
        return created, skipped

    def _import_evaluations(
        self, path: Path, *, dry_run: bool, build_reports: bool
    ) -> tuple[int, int]:
        User = get_user_model()
        created = skipped = 0
        for raw in _load_rows(path):
            phone = str(raw.get("شماره تماس") or "").strip()
            if not phone:
                skipped += 1
                continue
            created_raw = raw.get("تاریخ ثبت")
            created_at = _parse_dt(created_raw) if created_raw else None
            dup_qs = EvaluationRequest.objects.filter(phone=phone)
            if created_at:
                dup_qs = dup_qs.filter(created_at=created_at)
            if dup_qs.exists():
                self.stdout.write(
                    self.style.WARNING(f"Skip evaluation phone={phone} at {created_raw}")
                )
                skipped += 1
                continue

            payload = _map_row(raw, EVALUATION_COLUMNS)
            timestamps = {
                k: payload.pop(k)
                for k in (
                    "created_at",
                    "updated_at",
                    "admin_seen_at",
                    "contacted_at",
                    "next_follow_up_at",
                )
                if k in payload
            }
            assigned_id = payload.pop("assigned_to_id", None)
            snapshot = payload.pop("recommendation_snapshot", None)

            if assigned_id and not User.objects.filter(pk=assigned_id).exists():
                assigned_id = None

            if dry_run:
                self.stdout.write(f"Would create evaluation {payload.get('full_name')} ({phone})")
                created += 1
                continue

            obj = EvaluationRequest(
                **payload,
                recommendation_snapshot=snapshot,
                assigned_to_id=assigned_id,
            )
            obj.save()
            update_fields = {k: v for k, v in timestamps.items() if v is not None}
            if update_fields:
                EvaluationRequest.objects.filter(pk=obj.pk).update(**update_fields)
            obj.refresh_from_db()
            if build_reports:
                if snapshot and isinstance(snapshot, dict) and snapshot.get("has_data"):
                    create_evaluation_share(obj, report=snapshot)
                else:
                    create_evaluation_share(obj)
            created += 1
            self.stdout.write(self.style.SUCCESS(f"Evaluation pk={obj.pk} phone={phone}"))
        return created, skipped
