"""Analyze GSC Excel exports — dump all sheets to JSON for SEO review."""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
FILES = [
    ROOT / "saroshan.ir-Coverage-2026-05-26.xlsx",
    ROOT / "saroshan.ir-Performance-on-Search-2026-05-26.xlsx",
]
OUT = ROOT / "scripts" / "gsc_analysis_output.json"


def sheet_to_rows(ws) -> list[list]:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([("" if c is None else c) for c in row])
    return rows


def main() -> None:
    report: dict = {}
    for path in FILES:
        if not path.exists():
            report[path.name] = {"error": "not found"}
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        book: dict = {"sheets": {}}
        for name in wb.sheetnames:
            ws = wb[name]
            rows = sheet_to_rows(ws)
            book["sheets"][name] = {
                "row_count": len(rows),
                "headers": rows[0] if rows else [],
                "rows": rows[1:],
            }
        wb.close()
        report[path.name] = book

    OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {OUT}")
    for fname, book in report.items():
        if "error" in book:
            print(fname, book["error"])
            continue
        print(f"\n{fname}:")
        for sname, sdata in book["sheets"].items():
            print(f"  {sname}: {sdata['row_count']} rows")


if __name__ == "__main__":
    main()
    sys.exit(0)
